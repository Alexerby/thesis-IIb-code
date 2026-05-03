import json
from pathlib import Path
from typing import cast
import pandas as pd
import pyarrow.parquet as pq
from openpyxl.styles import Alignment, Font, PatternFill


def load_config(root_dir: str = ".") -> dict:
    """
    Load the project configuration from ``config.json``.

    Parameters
    ----------
    root_dir : str, optional
        Directory containing ``config.json``. Defaults to the current
        working directory.

    Returns
    -------
    dict
        Parsed JSON configuration.
    """
    config_path = Path(root_dir) / "config.json"
    with open(config_path) as f:
        return json.load(f)


def load_parquet_datasets(
    config: dict, parquet_dir: str = "output/data"
) -> tuple[list[pd.DataFrame], list[pd.DataFrame]]:
    """
    Load all Parquet datasets declared in config variables.

    Datasets are split into person-level (keyed by ``pid`` + ``syear``) and
    household-level (keyed by ``hid`` + ``syear``), which are merged
    separately in the pipeline.

    Parameters
    ----------
    config : dict
        Project config with a ``"variables"`` block mapping panel keys to
        lists of variable definitions, each having ``dataset`` and ``name``
        fields.
    parquet_dir : str, optional
        Directory containing the Parquet files. Defaults to
        ``"output/data"``.

    Returns
    -------
    tuple[list[pd.DataFrame], list[pd.DataFrame]]
        A 2-tuple ``(person_frames, household_frames)`` where each element
        is a list of DataFrames ready for merging.

    Raises
    ------
    FileNotFoundError
        If a required Parquet file is missing (run ``extract.py`` first).
    """
    from src.data.utils import HOUSEHOLD_DATASETS

    parquet_path = Path(parquet_dir)
    by_dataset = {}
    for panel in config["variables"].values():
        for vdef in panel:
            by_dataset.setdefault(vdef["dataset"], set()).add(vdef["name"].lower())
    for ddef in config.get("derived", []):
        for src in ddef.get("source_vars", []):
            by_dataset.setdefault(src["dataset"], set()).add(src["name"].lower())

    person_frames, household_frames = [], []
    for dataset, col_names in by_dataset.items():
        if dataset == "derived":
            continue
        path = parquet_path / f"{dataset}.parquet"
        if not path.exists():
            raise FileNotFoundError(f"Missing {path} - run extract.py first")
        id_col = "hid" if dataset in HOUSEHOLD_DATASETS else "pid"
        wanted = {id_col, "syear"} | {c.lower() for c in col_names}
        available = set(pq.read_schema(path).names)
        load_cols = [
            c
            for c in [id_col, "syear"] + sorted(wanted - {id_col, "syear"})
            if c in available
        ]
        df = pd.read_parquet(path, columns=load_cols)
        if dataset in HOUSEHOLD_DATASETS:
            household_frames.append(df)
        else:
            person_frames.append(df)
    return person_frames, household_frames


def _build_label_map(config: dict) -> dict[str, str]:
    """
    Build a flat variable-name → label mapping from all config sections.

    Covers direct variables, derived variables, and harmonized variables.
    A fallback label for ``syear`` is included since it is an implicit
    merge key not listed in ``config["variables"]``.

    Parameters
    ----------
    config : dict
        Full project configuration.

    Returns
    -------
    dict[str, str]
        Mapping of lowercase variable name → human-readable label.
    """
    label_map: dict[str, str] = {"syear": "Survey Year"}
    for panel in config["variables"].values():
        for vdef in panel:
            label_map[vdef["name"].lower()] = vdef["label"]
    for ddef in config.get("derived", []):
        label_map[ddef["name"].lower()] = ddef["label"]
    for hdef in config.get("harmonize", []):
        label_map[hdef["name"].lower()] = hdef["label"]
    return label_map


def _write_labeled_sheet(
    df: pd.DataFrame,
    writer: pd.ExcelWriter,
    sheet_name: str,
    label_map: dict[str, str],
) -> None:
    """
    Write ``df`` to an Excel sheet with variable names in row 1 and labels in row 2.

    pandas does not support ``index=False`` with MultiIndex columns, so this
    writes the data normally (row 1 = variable names), then uses openpyxl to
    insert a label row at position 2 and applies header styling.

    Parameters
    ----------
    df : pd.DataFrame
        DataFrame to write (flat column names).
    writer : pd.ExcelWriter
        Open ExcelWriter targeting an openpyxl workbook.
    sheet_name : str
        Name of the sheet to create.
    label_map : dict[str, str]
        Mapping of lowercase column name → human-readable label, as
        returned by ``_build_label_map``.
    """
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]

    ws.insert_rows(2)
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=2, column=col_idx).value = label_map.get(col_name, "")

    navy = PatternFill("solid", fgColor="1F3864")
    smoke = PatternFill("solid", fgColor="ECECEC")
    center = Alignment(horizontal="center", vertical="center")
    wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = navy
        cell.alignment = center

    for cell in ws[2]:
        cell.font = Font(italic=True, size=9, color="555555")
        cell.fill = smoke
        cell.alignment = wrap_center

    ws.row_dimensions[2].height = 30
    ws.freeze_panes = "A3"

    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=8,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)


def save_master(
    df: pd.DataFrame,
    output_path: str = "output/data/master.parquet",
    config: dict | None = None,
) -> pd.DataFrame:
    """
    Persist the master dataframe to Parquet and a multi-sheet Excel workbook.

    When ``config`` is supplied, each Excel sheet gets a two-row header:
    row 1 shows variable names (bold, blue) and row 2 shows human-readable
    labels from config (italic, grey). Panes are frozen at row 3. The
    Parquet file always uses flat column names so downstream code is
    unaffected.

    Three Excel sheets are written:

    - ``Full`` — entire dataframe.
    - ``Complete cases`` — rows with no missing values.
    - ``Eligible`` — rows where ``plb0097`` is 1 or 2 (excludes "Not possible").

    Parameters
    ----------
    df : pd.DataFrame
        Master dataframe to save.
    output_path : str, optional
        Destination path for the Parquet file. The Excel file is written
        to the same location with a ``.xlsx`` extension. Defaults to
        ``"output/data/master.parquet"``.
    config : dict or None, optional
        Project config used to resolve variable labels for the Excel
        header. If ``None``, the Excel header shows variable names only
        (single row, no labels).

    Returns
    -------
    pd.DataFrame
        The input dataframe unchanged, enabling use in a ``.pipe()`` chain.
    """
    out_path = Path(output_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_parquet(out_path, index=False)

    complete = cast(pd.DataFrame, df.dropna())
    eligible = cast(pd.DataFrame, df[df["plb0097"].isin([0, 1])].dropna())  # Filter out plb0097=3

    xlsx_path = out_path.with_suffix(".xlsx")
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        if config is not None:
            label_map = _build_label_map(config)
            _write_labeled_sheet(df,       writer, "Full",           label_map)
            _write_labeled_sheet(complete, writer, "Complete cases", label_map)
            _write_labeled_sheet(eligible, writer, "Eligible",       label_map)
        else:
            df.to_excel(writer,       sheet_name="Full",           index=False)
            complete.to_excel(writer, sheet_name="Complete cases", index=False)
            eligible.to_excel(writer, sheet_name="Eligible",       index=False)

    print(
        f"\nFinal Master dataframe: {len(df):,} rows ({len(complete):,} complete cases)"
    )
    print(f"Wrote {out_path} and {xlsx_path.name}")
    return df
